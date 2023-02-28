import time
from openpyxl.reader.excel import load_workbook


#######
# CONSTS
######

input_file = "zeitbogen.xlsx"
output_file = "output.xlsx"

name = "Klausius Mausius"
hours_each_week = 11
days_each_week = 5
department = "CLOWN SCHOOL"


def input_header_data(sheet):
    sheet.cell(row=3, column=5).value = name
    sheet.cell(row=3, column=10).value = department
    sheet.cell(row=4, column=2).value = hours_each_week
    sheet.cell(row=4, column=5).value = days_each_week
    sheet.cell(row=4, column=7).value = time.strftime("%H:%M:%S", time.gmtime(3600 * hours_each_week / days_each_week)) 
    print(sheet.cell(4, 7).value)


def main():
    wb = load_workbook(input_file)
    for sheet in wb.worksheets:
        input_header_data(sheet)
    wb.save(output_file)


if __name__ == "__main__":
    main()

