import time
import random
import datetime

import numpy as np
from openpyxl.reader.excel import load_workbook

from conf import *

#####
# FUNCS
#####

def read_data():
    output = []

    with open(raw_data_input_file, 'r') as f:
        for line in f.readlines():
            line = line.split('#')[0]
            output.append(line.split(','))
    return output


def split_data():
    raw_data = read_data()
    splitted_data = []

    for i in range(12):
        splitted_data.append([])
        while len(raw_data) > 0 and int(raw_data[0][0]) == i + 1:
            splitted_data[-1].append(raw_data[0])
            raw_data.pop(0)
    return splitted_data


def get_time_for_date():
    seconds = 3600 * (hours_each_month / 4 / days_each_week + 0.5)
    time_delta = datetime.timedelta(seconds=seconds)
    
    hours, remainder = divmod(time_delta.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    return datetime.datetime(1904, 1, 1, hours, minutes, seconds)


def set_header_data(sheet, month):
    sheet.cell(3, 2).value = datetime.date(year, month, 1)
    sheet.cell(3, 5).value = name
    sheet.cell(4, 2).value = hours_each_month
    sheet.cell(4, 5).value = department


def set_hour(sheet, hour, cell):
    if hour <= hours_threshold:
        return

    sheet.cell(cell[0], cell[1]).value = format_time(hour)


def set_hours(sheet, data):
    for d in data:
        cell = (6 + int(d[1]), 3)
        hour = round(float(d[2]) * round_hour_by) / round_hour_by
        set_hour(sheet, hour, cell)
        

def convert_info_to_dict(data) -> dict:
    hours = [sum([float(h[2]) for h in x]) for x in data]
    monthly_hours_unrounded = [sum(float(h[2]) for h in x) for x in data]
    monthly_hours = [round(x) for x in monthly_hours_unrounded]
    info = {
        "total_hours": round(sum(hours)),
        "monthly_hours": monthly_hours,
        "expected_month_hour": hours_each_month
    }
    return info


def main() -> dict:
    data = split_data()
    wb = load_workbook(xlsx_input_file)

    for index, sheet in enumerate(wb.worksheets):
        set_header_data(sheet, index + 1)
        set_hours(sheet, data[index])
    wb.save(xlsx_output_file)

    return convert_info_to_dict(data)


if __name__ == "__main__":
    info = main()

    print(info)
    print("Xlsx generated\n\n")

